#!/usr/bin/env python3

from typing import Tuple
from openpyxl import load_workbook
from openpyxl import Workbook


class Acp:

    @staticmethod
    def load_table(filepath: str) -> Tuple[list, list]:
        wb = load_workbook(filepath)
        sheet = wb.active
        title = []
        raw_table = []
        first_row = True
        for row in sheet:
            if first_row:
                for r in row:
                    title.append(str(r.value).strip())
                first_row = False
                continue
            row_a = [str(v.value).strip() for v in row]
            raw_table.append(row_a)
        return title, raw_table

    def group_by_npp(self, raw_table: list, entry_npp: int, entry_bs: int) -> dict:
        tb = {}
        for rows in raw_table:
            work_rows = self.string_to_float_item(rows, entry_bs)
            tb.setdefault(rows[entry_npp], []).append(work_rows)
        return tb

    @staticmethod
    def string_to_float_item(rows: list, entry_bs: int) -> list:
        m_rows = []
        i = 0
        while i < len(rows):
            if i == entry_bs:
                m_rows.append(float(rows[i].replace(',', '.')))
            else:
                m_rows.append(rows[i])
            i += 1
        return m_rows

    @staticmethod
    def number_of_unique_companies(tabl_by_npp: dict, enrty_kd: int) -> dict:
        tb = {}
        for npp, rows in tabl_by_npp.items():
            kds = []
            for row in rows:
                kds.append(row[enrty_kd])
            tb[npp] = len(set(kds))
        return tb

    @staticmethod
    def specific_weight(tabl_by_npp: dict, entry_kd: int, entry_bs: int) -> dict:
        sp_weight = {}
        for npp, rows in tabl_by_npp.items():
            kds = []
            npp_sum = 0
            for row in rows:
                kds.append(row[entry_kd])
                npp_sum = npp_sum + float(row[entry_bs])
            for uniq_kd in set(kds):
                kd_sum = 0
                for row in rows:
                    if uniq_kd == row[entry_kd]:
                        kd_sum = kd_sum + float(row[entry_bs])
                kd_percent = kd_sum / npp_sum * 100
                sp_weight.setdefault(npp, {}).setdefault(uniq_kd, []).extend([kd_sum, npp_sum, round(kd_percent, 2)])
        return sp_weight

    @staticmethod
    def write_to_file_new(file_name: str, table: dict, table_count: dict):
        wb = Workbook()
        wb.save(file_name)
        wb = load_workbook(file_name)
        sheet = wb["Sheet"]
        sheet.title = "Specific weight"
        for level in table.keys():
            for market in table[level].keys():
                sheet = wb.create_sheet(level + " " + market)
                for col_id, data in enumerate(['SORT', 'KD', 'BS_KD', 'BS', 'PERCENT_KD'], start=1):
                    sheet.cell(row=1, column=col_id).value = data
                i = 2
                for npp, rows in table[level][market].items():
                    for kd, row in rows.items():
                        sheet.cell(row=i, column=1).value = npp
                        sheet.cell(row=i, column=2).value = kd
                        sheet.cell(row=i, column=3).value = row[0]
                        sheet.cell(row=i, column=4).value = row[1]
                        sheet.cell(row=i, column=5).value = row[2]
                        i += 1
        wb.remove_sheet(wb.get_sheet_by_name("Specific weight"))

        sheet_2 = wb.create_sheet("Count kd")

        for col_id, data in enumerate(['SORT', 'COUNT_KD'], start=1):
            sheet_2.cell(row=1, column=col_id).value = data
        i = 2
        for npp, count in table_count.items():
            sheet_2.cell(row=i, column=1).value = npp
            sheet_2.cell(row=i, column=2).value = count
            i += 1
        wb.save(file_name)
